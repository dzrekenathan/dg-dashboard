import io
import uuid
from datetime import datetime, timezone
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
from openpyxl import load_workbook, Workbook
from app.database import get_db
from app.models import Task, User
from app.schemas import TaskOut, TaskUpdate
from app.core.security import get_current_user, require_management
from app.core.ws_manager import manager

router = APIRouter(prefix="/tasks", tags=["tasks"])

VALID_STATUSES = {"Not Started", "In Progress", "Completed", "On Hold", "At Risk", "Cancelled"}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _row_to_task(row: dict, fallback_so: Optional[str] = None) -> dict:
    """Map a spreadsheet row dict to a Task field dict."""
    so = str(
        row.get("SO #") or row.get("SO#") or row.get("so_number") or fallback_so or ""
    ).strip()
    status = str(row.get("Status") or row.get("status") or "Not Started").strip()
    if status not in VALID_STATUSES:
        status = "Not Started"
    try:
        progress = int(float(row.get("Progress %") or row.get("progress_pct") or 0))
    except (ValueError, TypeError):
        progress = 0
    return {
        "id": str(uuid.uuid4()),
        "so_number": so,
        "so_title": str(row.get("Strategic Objective") or row.get("so_title") or ""),
        "thematic_area": str(row.get("Thematic Area") or row.get("thematic_area") or ""),
        "task": str(row.get("Task") or row.get("task") or ""),
        "reference_numbers": str(row.get("Reference Nos.") or row.get("Reference No.") or row.get("reference_numbers") or ""),
        "activities": str(row.get("Activities & Sub-Activities") or row.get("Activities") or row.get("activities") or ""),
        "timeframe": str(row.get("Timeframe") or row.get("timeframe") or ""),
        "responsibility": str(row.get("Responsibility") or row.get("responsibility") or ""),
        "outputs": str(row.get("Outputs / Deliverables") or row.get("Outputs") or row.get("outputs") or ""),
        "outcomes": str(row.get("Outcomes / Impact") or row.get("Outcomes") or row.get("outcomes") or ""),
        "risks_mitigation": str(row.get("Risks & Mitigation") or row.get("Risks") or row.get("risks_mitigation") or ""),
        "budget": str(row.get("Budget") or row.get("budget") or ""),
        "status": status,
        "progress_pct": progress,
        "assigned_to": str(row.get("Assigned To") or row.get("assigned_to") or ""),
        "target_date": str(row.get("Target Date") or row.get("target_date") or ""),
        "notes": str(row.get("Notes / Comments") or row.get("Notes") or row.get("notes") or ""),
        "last_updated": "",
        "updated_by": "",
    }


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("", response_model=list[TaskOut])
async def list_tasks(
    so_number: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    stmt = select(Task)
    if so_number:
        stmt = stmt.where(Task.so_number == so_number)
    result = await db.execute(stmt.order_by(Task.so_number, Task.id))
    return result.scalars().all()


@router.patch("/{task_id}", response_model=TaskOut)
async def update_task(
    task_id: str,
    body: TaskUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_management),
):
    result = await db.execute(select(Task).where(Task.id == task_id))
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    for field, value in body.model_dump(exclude_none=True).items():
        setattr(task, field, value)

    task.last_updated = datetime.now(timezone.utc).isoformat()
    task.updated_by = body.updated_by or user.name

    await db.commit()
    await db.refresh(task)

    # Broadcast update to all connected clients
    await manager.broadcast({"type": "TASK_UPDATED", "payload": {"id": task.id, "so_number": task.so_number}})

    return task


@router.post("/import", response_model=dict)
async def import_tasks(
    file: UploadFile = File(...),
    so_number: Optional[str] = Query(None, description="If set, only replace tasks for this SO"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_management),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xls files are accepted")

    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb["SO Matrix"] if "SO Matrix" in wb.sheetnames else wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: (row[i] if row[i] is not None else "") for i in range(len(headers))}
        so = str(row_dict.get("SO #") or row_dict.get("SO#") or row_dict.get("so_number") or "").strip()
        if not so:
            continue
        if so_number and so != so_number:
            continue
        rows.append(_row_to_task(row_dict))

    if not rows:
        raise HTTPException(status_code=422, detail="No valid rows found in the file")

    # Delete existing tasks for affected SOs
    affected_sos = {r["so_number"] for r in rows}
    await db.execute(delete(Task).where(Task.so_number.in_(affected_sos)))

    # Insert new tasks
    db.add_all([Task(**r) for r in rows])
    await db.commit()

    await manager.broadcast({"type": "TASKS_UPDATED", "payload": {"so_numbers": list(affected_sos)}})

    return {"imported": len(rows), "so_numbers": list(affected_sos)}


@router.get("/export")
async def export_tasks(
    so_number: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    stmt = select(Task)
    if so_number:
        stmt = stmt.where(Task.so_number == so_number)
    result = await db.execute(stmt.order_by(Task.so_number, Task.id))
    tasks = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "SO Matrix"

    headers = [
        "SO #", "Strategic Objective", "Thematic Area", "Task", "Reference Nos.",
        "Activities & Sub-Activities", "Timeframe", "Responsibility",
        "Outputs / Deliverables", "Outcomes / Impact", "Risks & Mitigation",
        "Budget", "Status", "Progress %", "Assigned To", "Target Date",
        "Notes / Comments", "Last Updated", "Updated By",
    ]
    ws.append(headers)

    for t in tasks:
        ws.append([
            t.so_number, t.so_title, t.thematic_area, t.task, t.reference_numbers,
            t.activities, t.timeframe, t.responsibility, t.outputs, t.outcomes,
            t.risks_mitigation, t.budget, t.status, t.progress_pct, t.assigned_to,
            t.target_date, t.notes, t.last_updated, t.updated_by,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"CLET_SO_Matrix{'_' + so_number if so_number else ''}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
